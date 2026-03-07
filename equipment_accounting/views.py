import io
from datetime import date
from functools import wraps

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.contenttypes.models import ContentType
from django.core.exceptions import PermissionDenied
from django.db.models.deletion import ProtectedError
from django.core.paginator import Paginator
from django.db.models import Count, Exists, OuterRef, Prefetch, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse

from .forms import _get_available_uavs_for_kind
from .forms import (
    UAVInstanceForm, ComponentForm, PowerTemplateForm, VideoTemplateForm,
    FPVDroneTypeForm, OpticalDroneTypeForm,
    ManufacturerForm, DroneModelForm, LocationForm, PositionForm,
)
from .models import (
    UAVInstance, Component, PowerTemplate, VideoTemplate,
    FPVDroneType, OpticalDroneType,
    OtherComponentType, Location, UAVMovement,
    Manufacturer, DroneModel, UAVPhoto, DroneRole, Position,
)

def _list_url(tab="drones"):
    return reverse("equipment_accounting:equipment_list") + f"?tab={tab}"


GROUP_NAME = "майстер"
COMMANDER_GROUP = "командир майстерні"

# UAV permission codenames
PERM_ADD_UAV    = 'equipment_accounting.add_uavinstance'
PERM_CHANGE_UAV = 'equipment_accounting.change_uavinstance'
PERM_DELETE_UAV = 'equipment_accounting.delete_uavinstance'

PERM_ADD_COMPONENT    = 'equipment_accounting.add_component'
PERM_CHANGE_COMPONENT = 'equipment_accounting.change_component'
PERM_DELETE_COMPONENT = 'equipment_accounting.delete_component'

PERM_ADD_MANUFACTURER    = 'equipment_accounting.add_manufacturer'
PERM_CHANGE_MANUFACTURER = 'equipment_accounting.change_manufacturer'
PERM_DELETE_MANUFACTURER = 'equipment_accounting.delete_manufacturer'

PERM_ADD_DRONEMODEL    = 'equipment_accounting.add_dronemodel'
PERM_CHANGE_DRONEMODEL = 'equipment_accounting.change_dronemodel'
PERM_DELETE_DRONEMODEL = 'equipment_accounting.delete_dronemodel'

PERM_ADD_FPVTYPE    = 'equipment_accounting.add_fpvdronetype'
PERM_CHANGE_FPVTYPE = 'equipment_accounting.change_fpvdronetype'
PERM_DELETE_FPVTYPE = 'equipment_accounting.delete_fpvdronetype'

PERM_ADD_OPTICALTYPE    = 'equipment_accounting.add_opticaldronetype'
PERM_CHANGE_OPTICALTYPE = 'equipment_accounting.change_opticaldronetype'
PERM_DELETE_OPTICALTYPE = 'equipment_accounting.delete_opticaldronetype'

PERM_ADD_POWERTEMPLATE    = 'equipment_accounting.add_powertemplate'
PERM_CHANGE_POWERTEMPLATE = 'equipment_accounting.change_powertemplate'
PERM_DELETE_POWERTEMPLATE = 'equipment_accounting.delete_powertemplate'

PERM_ADD_VIDEOTEMPLATE    = 'equipment_accounting.add_videotemplate'
PERM_CHANGE_VIDEOTEMPLATE = 'equipment_accounting.change_videotemplate'
PERM_DELETE_VIDEOTEMPLATE = 'equipment_accounting.delete_videotemplate'

PERM_ADD_LOCATION    = 'equipment_accounting.add_location'
PERM_CHANGE_LOCATION = 'equipment_accounting.change_location'
PERM_DELETE_LOCATION = 'equipment_accounting.delete_location'

PERM_ADD_POSITION    = 'equipment_accounting.add_position'
PERM_CHANGE_POSITION = 'equipment_accounting.change_position'
PERM_DELETE_POSITION = 'equipment_accounting.delete_position'


def _is_master(user):
    """True for superusers and legacy master/commander group members."""
    return user.is_superuser or user.groups.filter(
        name__in=[GROUP_NAME, COMMANDER_GROUP]
    ).exists()


def _can(user, perm):
    """Check a specific permission, granting full access to master/commander users."""
    return _is_master(user) or user.has_perm(perm)


def master_required(view_func):
    """Allow access to superusers, master/commander groups, or any UAV-permission holder."""
    @wraps(view_func)
    @login_required
    def _wrapped(request, *args, **kwargs):
        if (_is_master(request.user)
                or request.user.has_perm('equipment_accounting.view_uavinstance')
                or request.user.has_perm(PERM_ADD_UAV)
                or request.user.has_perm(PERM_CHANGE_UAV)
                or request.user.has_perm(PERM_DELETE_UAV)):
            return view_func(request, *args, **kwargs)
        raise PermissionDenied
    return _wrapped


def uav_perm_required(perm):
    """Decorator: require a specific UAV permission (or master group membership)."""
    def decorator(view_func):
        @wraps(view_func)
        @login_required
        def _wrapped(request, *args, **kwargs):
            if _can(request.user, perm):
                return view_func(request, *args, **kwargs)
            raise PermissionDenied
        return _wrapped
    return decorator


# ── Main list view ──────────────────────────────────────────────────

@master_required
def equipment_list(request):
    tab = request.GET.get("tab", "drones")

    status_filter = request.GET.get("status", "")
    category_filter = request.GET.get("category", "")
    mode_filter = request.GET.get("mode", "")
    type_filter = request.GET.get("type", "")
    kit_filter = request.GET.get("kit", "")
    purpose_filter = request.GET.get("purpose", "")
    role_filter = request.GET.get("role", "")
    _location_raw = request.GET.get("location", "")
    location_filter = int(_location_raw) if _location_raw.isdigit() else None
    date_from = request.GET.get("date_from", "")
    date_to = request.GET.get("date_to", "")
    search_q = request.GET.get("q", "")

    # Compute ContentTypes once — reused by filters, group building, and type choices
    _fpv_ct = ContentType.objects.get_for_model(FPVDroneType)
    _opt_ct = ContentType.objects.get_for_model(OpticalDroneType)
    _fpv_ct_id = _fpv_ct.id
    _opt_ct_id = _opt_ct.id

    # Annotate with EXISTS subqueries for kit status — replaces prefetch_related on main queryset
    uavs = UAVInstance.objects.annotate(
        _has_battery=Exists(Component.objects.filter(assigned_to_uav=OuterRef('pk'), kind='battery')),
        _has_spool=Exists(Component.objects.filter(assigned_to_uav=OuterRef('pk'), kind='spool')),
    ).exclude(status='deleted')

    if location_filter:
        uavs = uavs.filter(
            Q(current_location_id=location_filter) |
            Q(status='transit', pending_to_location_id=location_filter)
        )

    if status_filter:
        uavs = uavs.filter(status=status_filter)

    if category_filter == "fpv":
        uavs = uavs.filter(content_type=_fpv_ct)
    elif category_filter == "optical":
        uavs = uavs.filter(content_type=_opt_ct)

    if type_filter:
        try:
            ct_id, obj_id = type_filter.split("-")
            uavs = uavs.filter(content_type_id=int(ct_id), object_id=int(obj_id))
        except (ValueError, TypeError):
            pass

    if date_from:
        try:
            uavs = uavs.filter(created_at__date__gte=date.fromisoformat(date_from))
        except ValueError:
            pass

    if date_to:
        try:
            uavs = uavs.filter(created_at__date__lte=date.fromisoformat(date_to))
        except ValueError:
            pass

    if search_q:
        fpv_ids = FPVDroneType.objects.filter(
            Q(model__name__icontains=search_q) | Q(model__manufacturer__name__icontains=search_q)
        ).values_list('pk', flat=True)
        opt_ids = OpticalDroneType.objects.filter(
            Q(model__name__icontains=search_q) | Q(model__manufacturer__name__icontains=search_q)
        ).values_list('pk', flat=True)
        uavs = uavs.filter(
            Q(notes__icontains=search_q) |
            Q(content_type=_fpv_ct, object_id__in=fpv_ids) |
            Q(content_type=_opt_ct, object_id__in=opt_ids)
        )

    # Kit filter at DB level via EXISTS annotations
    if kit_filter == UAVInstance.KIT_NONE:
        uavs = uavs.filter(_has_battery=False, _has_spool=False)
    elif kit_filter == UAVInstance.KIT_FULL:
        uavs = uavs.filter(
            Q(content_type=_opt_ct, _has_battery=True, _has_spool=True) |
            Q(_has_battery=True) & ~Q(content_type=_opt_ct)
        )
    elif kit_filter == UAVInstance.KIT_PARTIAL:
        uavs = uavs.exclude(
            Q(_has_battery=False, _has_spool=False) |
            Q(content_type=_opt_ct, _has_battery=True, _has_spool=True) |
            Q(_has_battery=True) & ~Q(content_type=_opt_ct)
        )

    # Role filter
    if role_filter.isdigit():
        uavs = uavs.filter(role_id=int(role_filter))

    # Mode filter (день/ніч based on has_thermal)
    if mode_filter in ("day", "night"):
        is_thermal = (mode_filter == "night")
        _thermal_fpv = list(FPVDroneType.objects.filter(has_thermal=is_thermal).values_list('pk', flat=True))
        _thermal_opt = list(OpticalDroneType.objects.filter(has_thermal=is_thermal).values_list('pk', flat=True))
        uavs = uavs.filter(
            Q(content_type=_fpv_ct, object_id__in=_thermal_fpv) |
            Q(content_type=_opt_ct, object_id__in=_thermal_opt)
        )

    # Pre-fetch all drone types into dicts — avoids N+1 GenericFK access
    _fpv_types = {dt.pk: dt for dt in FPVDroneType.objects.select_related(
        "model", "video_frequency",
    ).prefetch_related("control_frequencies").only(
        "id", "prop_size", "has_thermal",
        "video_frequency_id", "video_frequency__value", "video_frequency__unit",
        "model__name",
    )}
    _opt_types = {dt.pk: dt for dt in OpticalDroneType.objects.select_related(
        "model", "video_template",
    ).only(
        "id", "prop_size", "has_thermal",
        "video_template_id", "video_template__max_distance",
        "model__name",
    )}

    def _kit_from_ann(uav):
        """Compute kit status from EXISTS annotations — no extra DB queries."""
        if not uav._has_battery and not uav._has_spool:
            return UAVInstance.KIT_NONE
        if uav.content_type_id == _opt_ct_id:
            if uav._has_battery and uav._has_spool:
                return UAVInstance.KIT_FULL
            return UAVInstance.KIT_PARTIAL
        return UAVInstance.KIT_FULL if uav._has_battery else UAVInstance.KIT_PARTIAL

    # Light query: only fields needed for group-building (no heavy select_related, no component rows)
    uavs_light = uavs.select_related("role").only(
        'id', 'content_type_id', 'object_id', 'status', 'role_id', 'created_at'
    ).order_by('-created_at')

    # Build badge groups, tracking UAV PKs per group (no full UAV objects yet)
    _badge_seen = {}
    badge_groups = []
    _group_uav_ids = {}
    _status_display = dict(UAVInstance.STATUS_CHOICES)

    for _uav in uavs_light:
        _kit = _kit_from_ann(_uav)
        _key = (_uav.content_type_id, _uav.object_id, _uav.created_at.date(), _kit)
        if _key not in _badge_seen:
            _is_opt = _uav.content_type_id == _opt_ct_id
            _dt = (_opt_types if _is_opt else _fpv_types).get(_uav.object_id)
            _is_th = _dt.has_thermal if _dt else False
            if not _is_opt:
                _purpose = 'ніч' if _is_th else 'день'
                _purpose_label = 'Ніч' if _is_th else 'День'
            else:
                _purpose = 'ударні'
                _purpose_label = 'Ударні'
            _g = {
                '_key': _key,
                'type_label': _make_list_type_label(_dt, _is_opt),
                'category': 'Оптика' if _is_opt else 'Радіо',
                'mode_label': 'Ніч' if _is_th else 'День',
                'purpose': _purpose,
                'purpose_label': _purpose_label,
                'role_name': _uav.role.name if _uav.role_id else '—',
                'date': _uav.created_at.date(),
                'type_key': f"{_uav.content_type_id}-{_uav.object_id}",
                'date_str': _uav.created_at.date().isoformat(),
                'kit_status': _kit,
                'kit_label': UAVInstance.KIT_LABELS[_kit],
                'total': 0,
                'status_counts': {},
                'uavs': [],
            }
            _badge_seen[_key] = _g
            badge_groups.append(_g)
            _group_uav_ids[_key] = []
        _bg = _badge_seen[_key]
        _bg['total'] += 1
        _bg['status_counts'][_uav.status] = _bg['status_counts'].get(_uav.status, 0) + 1
        _group_uav_ids[_key].append(_uav.pk)

    for _g in badge_groups:
        _g['status_items'] = [
            (s, _status_display.get(s, s), c)
            for s, c in _g['status_counts'].items()
        ]
        _g['cnt_ready']      = _g['status_counts'].get('ready', 0)
        _g['cnt_inspection'] = _g['status_counts'].get('inspection', 0)
        _g['cnt_repair']     = _g['status_counts'].get('repair', 0)
        _g['cnt_deferred']   = _g['status_counts'].get('deferred', 0)
        _g['cnt_transit']    = _g['status_counts'].get('transit', 0)

    # Paginate at GROUP level — each page shows up to 20 drone-type groups
    total_uavs = sum(_g['total'] for _g in badge_groups)
    paginator = Paginator(badge_groups, 20)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)
    current_groups = list(page_obj)

    # Load full UAV details only for the current page's groups
    current_uav_ids = [pk for _g in current_groups for pk in _group_uav_ids.get(_g['_key'], [])]
    if current_uav_ids:
        uavs_detail = {
            uav.pk: uav
            for uav in UAVInstance.objects.filter(pk__in=current_uav_ids)
                .select_related("content_type", "current_location", "position", "role", "pending_to_location")
                .prefetch_related(Prefetch("components", queryset=Component.objects.only("kind", "assigned_to_uav_id")))
        }
        for _g in current_groups:
            _g['uavs'] = [uavs_detail[pk] for pk in _group_uav_ids.get(_g['_key'], []) if pk in uavs_detail]

    # Build drone type choices — reuse already-fetched type dicts (no extra queries)
    type_choices = (
        [(f"{_fpv_ct_id}-{dt.pk}", f"[Радіо] {dt}") for dt in sorted(_fpv_types.values(), key=str)] +
        [(f"{_opt_ct_id}-{dt.pk}", f"[Оптика] {dt}") for dt in sorted(_opt_types.values(), key=str)]
    )

    # Summary counts — single aggregated query, total derived from it (no COUNT(*) needed)
    all_uavs = UAVInstance.objects.exclude(status='deleted')
    _status_agg = {row['status']: row['cnt'] for row in all_uavs.values('status').annotate(cnt=Count('pk'))}
    total_drones = sum(_status_agg.values())
    status_counts = {
        code: {"label": label, "count": _status_agg.get(code, 0)}
        for code, label in UAVInstance.STATUS_CHOICES if code != 'deleted'
    }

    # Components with filters
    comp_status_filter     = request.GET.get("comp_status", "")
    comp_category_filter   = request.GET.get("comp_category", "")
    comp_assign_filter     = request.GET.get("comp_assign", "")
    comp_model_filter      = request.GET.get("comp_model", "")
    comp_drone_type_filter = request.GET.get("comp_drone_type", "")

    # Only query components when the components tab is active (avoids COUNT + scan on every drone-tab load)
    if tab == 'components':
        components_qs = Component.objects.select_related(
            "power_template", "video_template", "other_type", "assigned_to_uav"
        ).exclude(status='given').order_by("-created_at")
        if comp_status_filter:
            components_qs = components_qs.filter(status=comp_status_filter)
        if comp_category_filter in ('battery', 'spool', 'other'):
            components_qs = components_qs.filter(kind=comp_category_filter)
        if comp_assign_filter == "assigned":
            components_qs = components_qs.filter(assigned_to_uav__isnull=False)
        elif comp_assign_filter == "free":
            components_qs = components_qs.filter(assigned_to_uav__isnull=True)
        if comp_model_filter:
            try:
                model_id = int(comp_model_filter)
            except ValueError:
                model_id = None
            if model_id:
                pt_ids = list(
                    FPVDroneType.objects.filter(model_id=model_id).values_list('power_template_id', flat=True)
                ) + list(
                    OpticalDroneType.objects.filter(model_id=model_id).values_list('power_template_id', flat=True)
                )
                components_qs = components_qs.filter(
                    Q(kind='battery', power_template_id__in=pt_ids) |
                    Q(kind='spool', video_template__drone_model_id=model_id)
                )
        if comp_drone_type_filter == 'fpv':
            fpv_pt_ids = FPVDroneType.objects.values_list('power_template_id', flat=True)
            components_qs = components_qs.filter(kind='battery', power_template_id__in=fpv_pt_ids)
        elif comp_drone_type_filter == 'optical':
            opt_pt_ids = OpticalDroneType.objects.values_list('power_template_id', flat=True)
            components_qs = components_qs.filter(
                Q(kind='battery', power_template_id__in=opt_pt_ids) |
                Q(kind='spool')
            )
        comp_paginator = Paginator(components_qs, 20)
        comp_page_obj  = comp_paginator.get_page(request.GET.get("cpage"))
    else:
        comp_page_obj = None

    # Drone types
    fpv_drone_types = FPVDroneType.objects.select_related(
        "model", "model__manufacturer", "purpose",
        "video_frequency", "power_template",
    ).prefetch_related("control_frequencies")
    optical_drone_types = OpticalDroneType.objects.select_related(
        "model", "model__manufacturer", "purpose",
        "video_template", "video_template__drone_model", "power_template",
    ).prefetch_related("control_frequencies")

    # Templates (exclude soft-deleted)
    power_templates = PowerTemplate.objects.filter(is_deleted=False)
    video_templates = VideoTemplate.objects.filter(is_deleted=False).select_related("drone_model")

    # Reference data — evaluated once, reused in ctx to avoid duplicate queries
    manufacturers = Manufacturer.objects.all()
    drone_models = DroneModel.objects.select_related("manufacturer").all()
    _locations = list(Location.objects.annotate(uav_count=Count('current_uavs')))
    position_location_ids = [loc.pk for loc in _locations if loc.name == 'Позиція']
    _positions = list(Position.objects.annotate(uav_count=Count('uavs')))

    ctx = {
        "tab": tab,
        "page_obj": page_obj,
        "status_filter": status_filter,
        "category_filter": category_filter,
        "mode_filter": mode_filter,
        "type_filter": type_filter,
        "kit_filter": kit_filter,
        "kit_choices": list(UAVInstance.KIT_LABELS.items()),
        "purpose_filter": purpose_filter,
        "role_filter": role_filter,
        "drone_roles": DroneRole.objects.all(),
        "location_filter": location_filter,
        "date_from": date_from,
        "date_to": date_to,
        "search_q": search_q,
        "type_choices": type_choices,
        "total_drones": total_drones,
        "status_counts": status_counts,
        "status_choices": [c for c in UAVInstance.STATUS_CHOICES if c[0] != 'deleted'],
        "comp_page_obj": comp_page_obj,
        "comp_status_filter": comp_status_filter,
        "comp_category_filter": comp_category_filter,
        "comp_assign_filter": comp_assign_filter,
        "comp_model_filter": comp_model_filter,
        "comp_drone_type_filter": comp_drone_type_filter,
        "fpv_drone_types": fpv_drone_types,
        "optical_drone_types": optical_drone_types,
        "power_templates": power_templates,
        "video_templates": video_templates,
        "manufacturers": manufacturers,
        "drone_models": drone_models,
        "locations": _locations,
        "locations_give": _locations,
        "all_positions": _positions,
        "total_uavs": total_uavs,
        "position_location_ids": position_location_ids,
        "can_add_uav":    _can(request.user, PERM_ADD_UAV),
        "can_edit_uav":   _can(request.user, PERM_CHANGE_UAV),
        "can_delete_uav": _can(request.user, PERM_DELETE_UAV),
        "positions": _positions,
        "can_add_component":    _is_master(request.user) or request.user.has_perm('equipment_accounting.add_component'),
        "can_edit_component":   _is_master(request.user) or request.user.has_perm('equipment_accounting.change_component'),
        "can_delete_component": _is_master(request.user) or request.user.has_perm('equipment_accounting.delete_component'),
        "can_add_manufacturer":    _can(request.user, PERM_ADD_MANUFACTURER),
        "can_edit_manufacturer":   _can(request.user, PERM_CHANGE_MANUFACTURER),
        "can_delete_manufacturer": _can(request.user, PERM_DELETE_MANUFACTURER),
        "can_add_dronemodel":    _can(request.user, PERM_ADD_DRONEMODEL),
        "can_edit_dronemodel":   _can(request.user, PERM_CHANGE_DRONEMODEL),
        "can_delete_dronemodel": _can(request.user, PERM_DELETE_DRONEMODEL),
        "can_add_fpvtype":    _can(request.user, PERM_ADD_FPVTYPE),
        "can_edit_fpvtype":   _can(request.user, PERM_CHANGE_FPVTYPE),
        "can_delete_fpvtype": _can(request.user, PERM_DELETE_FPVTYPE),
        "can_add_opticaltype":    _can(request.user, PERM_ADD_OPTICALTYPE),
        "can_edit_opticaltype":   _can(request.user, PERM_CHANGE_OPTICALTYPE),
        "can_delete_opticaltype": _can(request.user, PERM_DELETE_OPTICALTYPE),
        "can_add_powertemplate":    _can(request.user, PERM_ADD_POWERTEMPLATE),
        "can_edit_powertemplate":   _can(request.user, PERM_CHANGE_POWERTEMPLATE),
        "can_delete_powertemplate": _can(request.user, PERM_DELETE_POWERTEMPLATE),
        "can_see_powertemplate": _is_master(request.user) or any(
            request.user.has_perm(p) for p in [
                'equipment_accounting.view_powertemplate', 'equipment_accounting.add_powertemplate',
                'equipment_accounting.change_powertemplate', 'equipment_accounting.delete_powertemplate',
            ]
        ),
        "can_add_videotemplate":    _can(request.user, PERM_ADD_VIDEOTEMPLATE),
        "can_edit_videotemplate":   _can(request.user, PERM_CHANGE_VIDEOTEMPLATE),
        "can_delete_videotemplate": _can(request.user, PERM_DELETE_VIDEOTEMPLATE),
        "can_see_videotemplate": _is_master(request.user) or any(
            request.user.has_perm(p) for p in [
                'equipment_accounting.view_videotemplate', 'equipment_accounting.add_videotemplate',
                'equipment_accounting.change_videotemplate', 'equipment_accounting.delete_videotemplate',
            ]
        ),
        "can_tab_components": _is_master(request.user) or any(
            request.user.has_perm(p) for p in [
                'equipment_accounting.add_component',
                'equipment_accounting.change_component',
                'equipment_accounting.delete_component',
            ]
        ),
        "can_tab_types": _is_master(request.user) or any(
            request.user.has_perm(p) for p in [
                'equipment_accounting.view_manufacturer',  'equipment_accounting.add_manufacturer',    'equipment_accounting.change_manufacturer',    'equipment_accounting.delete_manufacturer',
                'equipment_accounting.view_dronemodel',   'equipment_accounting.add_dronemodel',       'equipment_accounting.change_dronemodel',      'equipment_accounting.delete_dronemodel',
                'equipment_accounting.view_fpvdronetype', 'equipment_accounting.add_fpvdronetype',     'equipment_accounting.change_fpvdronetype',    'equipment_accounting.delete_fpvdronetype',
                'equipment_accounting.view_opticaldronetype', 'equipment_accounting.add_opticaldronetype', 'equipment_accounting.change_opticaldronetype', 'equipment_accounting.delete_opticaldronetype',
            ]
        ),
        "can_tab_templates": _is_master(request.user) or any(
            request.user.has_perm(p) for p in [
                'equipment_accounting.view_powertemplate', 'equipment_accounting.add_powertemplate',  'equipment_accounting.change_powertemplate',  'equipment_accounting.delete_powertemplate',
                'equipment_accounting.view_videotemplate', 'equipment_accounting.add_videotemplate',  'equipment_accounting.change_videotemplate',  'equipment_accounting.delete_videotemplate',
            ]
        ),
        "can_add_location":    _can(request.user, PERM_ADD_LOCATION),
        "can_edit_location":   _can(request.user, PERM_CHANGE_LOCATION),
        "can_delete_location": _can(request.user, PERM_DELETE_LOCATION),
        "can_add_position":    _can(request.user, PERM_ADD_POSITION),
        "can_edit_position":   _can(request.user, PERM_CHANGE_POSITION),
        "can_delete_position": _can(request.user, PERM_DELETE_POSITION),
        "can_tab_locations": _is_master(request.user) or any(
            request.user.has_perm(p) for p in [
                'equipment_accounting.view_location',  'equipment_accounting.add_location',
                'equipment_accounting.change_location', 'equipment_accounting.delete_location',
                'equipment_accounting.view_position',  'equipment_accounting.add_position',
                'equipment_accounting.change_position', 'equipment_accounting.delete_position',
            ]
        ),
    }

    return render(request, "equipment_accounting/equipment_list.html", ctx)


# ── Component Statistics ─────────────────────────────────────────────

@master_required
def component_stats(request):
    battery_stats = PowerTemplate.objects.filter(is_deleted=False).annotate(
        total=Count('battery_components',
            filter=~Q(battery_components__status='given')),
        cnt_in_use=Count('battery_components',
            filter=Q(battery_components__status='in_use')),
        cnt_free=Count('battery_components',
            filter=Q(battery_components__status='disassembled')),
        cnt_damaged=Count('battery_components',
            filter=Q(battery_components__status='damaged')),
    ).filter(total__gt=0).order_by('name')

    spool_stats = VideoTemplate.objects.filter(is_deleted=False).prefetch_related(
        Prefetch('opticaldronetype_set', queryset=OpticalDroneType.objects.select_related('model'))
    ).annotate(
        total=Count('spool_components',
            filter=~Q(spool_components__status='given')),
        cnt_in_use=Count('spool_components',
            filter=Q(spool_components__status='in_use')),
        cnt_free=Count('spool_components',
            filter=Q(spool_components__status='disassembled')),
        cnt_damaged=Count('spool_components',
            filter=Q(spool_components__status='damaged')),
    ).filter(total__gt=0).order_by('name')

    other_stats = OtherComponentType.objects.annotate(
        total=Count('components',
            filter=~Q(components__status='given')),
        cnt_in_use=Count('components',
            filter=Q(components__status='in_use')),
        cnt_free=Count('components',
            filter=Q(components__status='disassembled')),
        cnt_damaged=Count('components',
            filter=Q(components__status='damaged')),
    ).filter(total__gt=0).order_by('category', 'model')

    def _kind_summary(kind):
        qs = Component.objects.filter(kind=kind).exclude(status='given')
        return {
            'total':   qs.count(),
            'in_use':  qs.filter(status='in_use').count(),
            'free':    qs.filter(status='disassembled').count(),
            'damaged': qs.filter(status='damaged').count(),
        }

    return render(request, 'equipment_accounting/component_stats.html', {
        'battery_stats':   battery_stats,
        'spool_stats':     spool_stats,
        'other_stats':     other_stats,
        'battery_summary': _kind_summary('battery'),
        'spool_summary':   _kind_summary('spool'),
    })


# ── Drone location statistics ─────────────────────────────────────────

@master_required
def drone_location_stats(request):
    """Show drone counts grouped by current location and status."""
    locations = Location.objects.annotate(
        total=Count('current_uavs', filter=~Q(current_uavs__status='deleted')),
        cnt_ready=Count('current_uavs', filter=Q(current_uavs__status='ready')),
        cnt_inspection=Count('current_uavs', filter=Q(current_uavs__status='inspection')),
        cnt_repair=Count('current_uavs', filter=Q(current_uavs__status='repair')),
        cnt_deferred=Count('current_uavs', filter=Q(current_uavs__status='deferred')),
        cnt_given=Count('current_uavs', filter=Q(current_uavs__status='given')),
    ).order_by('name')

    total_all = UAVInstance.objects.exclude(status='deleted').count()
    transit_total = UAVInstance.objects.filter(status='transit').count()

    # Per-position-name breakdown for position-type locations
    # (includes both confirmed drones and transit drones en route)
    _pos_sub = {}  # loc_id -> {position_name -> count}

    for item in (
        UAVInstance.objects
        .exclude(status__in=['deleted', 'transit'])
        .filter(current_location__name='Позиція')
        .values('current_location_id', 'position__name')
        .annotate(n=Count('pk'))
    ):
        lid = item['current_location_id']
        pname = item['position__name'] or ''
        _pos_sub.setdefault(lid, {})
        _pos_sub[lid][pname] = _pos_sub[lid].get(pname, 0) + item['n']

    for item in (
        UAVInstance.objects
        .filter(status='transit',
                pending_to_location__isnull=False,
                pending_to_location__name='Позиція')
        .values('pending_to_location_id', 'position__name')
        .annotate(n=Count('pk'))
    ):
        lid = item['pending_to_location_id']
        pname = item['position__name'] or ''
        _pos_sub.setdefault(lid, {})
        _pos_sub[lid][pname] = _pos_sub[lid].get(pname, 0) + item['n']

    pos_sub_rows = {
        lid: sorted(
            [{'name': k or '— без назви —', 'total': v} for k, v in d.items()],
            key=lambda x: x['name']
        )
        for lid, d in _pos_sub.items()
    }

    locations_with_pos = [
        {
            'loc': loc,
            'pos_rows': pos_sub_rows.get(loc.pk, []),
        }
        for loc in locations
    ]

    return render(request, 'equipment_accounting/drone_location_stats.html', {
        'locations': locations,
        'locations_with_pos': locations_with_pos,
        'total_all': total_all,
        'transit_total': transit_total,
    })


@master_required
def drone_stats(request):
    """Detailed drone count breakdown by type, mode, and status — with filters and Excel export."""
    _fpv_ct = ContentType.objects.get_for_model(FPVDroneType)
    _opt_ct = ContentType.objects.get_for_model(OpticalDroneType)

    ALL_STATUSES = [
        ('ready',      'Готовий'),
        ('inspection', 'Перевірка'),
        ('repair',     'Ремонт'),
        ('deferred',   'Відкладено'),
        ('transit',    'В дорозі'),
        ('given',      'Віддано'),
    ]
    ALL_STATUS_KEYS = [s for s, _ in ALL_STATUSES]

    fpv_types_map = {dt.pk: dt for dt in FPVDroneType.objects.select_related(
        'model', 'video_frequency',
    ).prefetch_related('control_frequencies')}
    opt_types_map = {dt.pk: dt for dt in OpticalDroneType.objects.select_related(
        'model', 'video_template',
    )}
    all_locations = list(Location.objects.all().order_by('name'))
    all_roles     = list(DroneRole.objects.all().order_by('name'))

    # ── Parse filters ────────────────────────────────────────────────
    # _f sentinel: if present the form was submitted; otherwise use defaults
    is_filtered = '_f' in request.GET

    if is_filtered:
        sel_loc_ids   = {int(x) for x in request.GET.getlist('loc')  if x.isdigit()}
        sel_stat_keys = [s for s in request.GET.getlist('stat') if s in ALL_STATUS_KEYS]
        sel_modes     = set(request.GET.getlist('mode'))   # 'day', 'night'
        sel_cats      = set(request.GET.getlist('cat'))    # 'fpv', 'optical'
        sel_role_ids  = {int(x) for x in request.GET.getlist('role') if x.isdigit()}
    else:
        sel_loc_ids   = {loc.pk for loc in all_locations}
        sel_stat_keys = list(ALL_STATUS_KEYS)
        sel_modes     = {'day', 'night'}
        sel_cats      = {'fpv', 'optical'}
        sel_role_ids  = {role.pk for role in all_roles}

    # Visible status columns
    statuses     = [(s, lbl) for s, lbl in ALL_STATUSES if s in sel_stat_keys]
    STATUS_KEYS  = [s for s, _ in statuses]

    # Location queryset filter
    all_loc_ids = {loc.pk for loc in all_locations}
    if sel_loc_ids and sel_loc_ids != all_loc_ids:
        loc_q = (Q(current_location_id__in=sel_loc_ids) |
                 Q(status='transit', pending_to_location_id__in=sel_loc_ids))
    else:
        loc_q = Q()   # no restriction

    # ── Helpers ──────────────────────────────────────────────────────
    def _tlabel(ct_id, obj_id):
        if ct_id == _fpv_ct.id:
            dt = fpv_types_map.get(obj_id)
            return _make_list_type_label(dt, False) if dt else f'FPV #{obj_id}'
        dt = opt_types_map.get(obj_id)
        return _make_list_type_label(dt, True) if dt else f'Opt #{obj_id}'

    def _build(type_q):
        raw = (
            UAVInstance.objects
            .filter(type_q & loc_q)
            .exclude(status='deleted')
            .values('content_type_id', 'object_id', 'status')
            .annotate(cnt=Count('pk'))
        )
        data = {}
        for row in raw:
            key = (row['content_type_id'], row['object_id'])
            data.setdefault(key, {s: 0 for s in ALL_STATUS_KEYS})
            if row['status'] in ALL_STATUS_KEYS:
                data[key][row['status']] = row['cnt']

        rows = []
        for key in sorted(data, key=lambda k: _tlabel(*k)):
            counts = data[key]
            total  = sum(counts[s] for s in STATUS_KEYS)
            rows.append({
                'name': _tlabel(*key),
                'status_counts': [(s, lbl, counts.get(s, 0)) for s, lbl in statuses],
                'total': total,
            })
        col_totals = {s: sum(data[k].get(s, 0) for k in data) for s in STATUS_KEYS}
        grand  = sum(col_totals.values())
        totals = [(s, lbl, col_totals[s]) for s, lbl in statuses]
        return rows, totals, grand

    # ── Build sections ───────────────────────────────────────────────
    fpv_day_ids   = [pk for pk, dt in fpv_types_map.items() if not dt.has_thermal]
    fpv_night_ids = [pk for pk, dt in fpv_types_map.items() if dt.has_thermal]
    opt_ids       = list(opt_types_map.keys())

    ROLE_COLORS = ['sky', 'violet', 'rose', 'amber', 'emerald', 'indigo']
    sections = []

    if 'day' in sel_modes and 'fpv' in sel_cats:
        rows, tots, grand = _build(Q(content_type=_fpv_ct, object_id__in=fpv_day_ids))
        if rows:
            sections.append({'name': 'День',   'subtitle': 'FPV · без термальної камери',
                             'color': 'day',    'rows': rows, 'totals': tots, 'grand': grand})

    if 'night' in sel_modes and 'fpv' in sel_cats:
        rows, tots, grand = _build(Q(content_type=_fpv_ct, object_id__in=fpv_night_ids))
        if rows:
            sections.append({'name': 'Ніч',    'subtitle': 'FPV · термальна камера',
                             'color': 'night',  'rows': rows, 'totals': tots, 'grand': grand})

    if 'optical' in sel_cats:
        rows, tots, grand = _build(Q(content_type=_opt_ct, object_id__in=opt_ids))
        if rows:
            sections.append({'name': 'Оптика', 'subtitle': 'Оптичні БПЛА',
                             'color': 'optical', 'rows': rows, 'totals': tots, 'grand': grand})

    for i, role in enumerate(all_roles):
        if role.pk not in sel_role_ids:
            continue
        rows, tots, grand = _build(Q(role=role))
        if rows:
            sections.append({'name': role.name, 'subtitle': f'Роль: {role.name}',
                             'color': ROLE_COLORS[i % len(ROLE_COLORS)],
                             'rows': rows, 'totals': tots, 'grand': grand})

    # ── Summary cards ────────────────────────────────────────────────
    summary_qs = UAVInstance.objects.exclude(status='deleted')
    if loc_q != Q():
        summary_qs = summary_qs.filter(loc_q)
    total_by_status = {s: 0 for s in ALL_STATUS_KEYS}
    for row in summary_qs.values('status').annotate(cnt=Count('pk')):
        if row['status'] in ALL_STATUS_KEYS:
            total_by_status[row['status']] = row['cnt']
    total_all     = sum(total_by_status.values())
    summary_cards = [(s, lbl, total_by_status[s]) for s, lbl in ALL_STATUSES]

    # ── Excel export ─────────────────────────────────────────────────
    if request.GET.get('export') == 'xlsx':
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        STAT_COLORS = {
            'ready': '15803D', 'inspection': '1D4ED8', 'repair': 'B45309',
            'deferred': '94A3B8', 'transit': '0F766E', 'given': '475569',
        }
        SEC_STYLES = {
            'День':   ('FEF3C7', '92400E'), 'Ніч':    ('E0E7FF', '3730A3'),
            'Оптика': ('F3E8FF', '6B21A8'),
        }
        DEF_STYLE = ('F1F5F9', '1E293B')

        wb = Workbook()
        ws = wb.active
        ws.title = 'Деталізація БПЛА'

        headers = ['Тип БПЛА'] + [lbl for _, lbl in statuses] + ['Всього']
        hdr_fill = PatternFill(fill_type='solid', fgColor='D6DCE4')
        hdr_font = Font(bold=True, size=10, name='Calibri')
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = hdr_font; c.fill = hdr_fill
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[1].height = 26

        rn = 2
        for sec in sections:
            bg, fg = SEC_STYLES.get(sec['name'], DEF_STYLE)
            sec_fill = PatternFill(fill_type='solid', fgColor=bg)
            sec_font = Font(bold=True, size=10, name='Calibri', color=fg)
            # section header row
            for ci in range(1, len(headers) + 1):
                c = ws.cell(row=rn, column=ci,
                            value=sec['name'] if ci == 1 else None)
                c.font = sec_font; c.fill = sec_fill
            rn += 1
            # data rows
            for dr in sec['rows']:
                ws.cell(row=rn, column=1, value=dr['name']).font = Font(size=10, name='Calibri')
                for ci, (sk, _, cnt) in enumerate(dr['status_counts'], 2):
                    c = ws.cell(row=rn, column=ci, value=cnt or None)
                    c.alignment = Alignment(horizontal='center')
                    if cnt and sk in STAT_COLORS:
                        c.font = Font(size=10, name='Calibri', color=STAT_COLORS[sk], bold=True)
                tot_c = ws.cell(row=rn, column=len(headers), value=dr['total'])
                tot_c.font = Font(bold=True, size=10, name='Calibri')
                tot_c.alignment = Alignment(horizontal='center')
                rn += 1
            # totals row
            tot_font = Font(bold=True, size=10, name='Calibri')
            ws.cell(row=rn, column=1, value='Всього').font = tot_font
            for ci, (_, _, cnt) in enumerate(sec['totals'], 2):
                c = ws.cell(row=rn, column=ci, value=cnt or None)
                c.font = tot_font; c.alignment = Alignment(horizontal='center')
            gc = ws.cell(row=rn, column=len(headers), value=sec['grand'])
            gc.font = tot_font; gc.alignment = Alignment(horizontal='center')
            rn += 2  # blank row between sections

        ws.column_dimensions[get_column_letter(1)].width = 38
        for ci in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 13

        resp = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="drone_stats.xlsx"'
        wb.save(resp)
        return resp

    # ── Render page ──────────────────────────────────────────────────
    filter_state = {
        'locations': [(loc, loc.pk in sel_loc_ids) for loc in all_locations],
        'statuses':  [(s, lbl, s in sel_stat_keys) for s, lbl in ALL_STATUSES],
        'modes':     [('day', 'День', 'day' in sel_modes),
                      ('night', 'Ніч', 'night' in sel_modes)],
        'cats':      [('fpv', 'Радіо (FPV)', 'fpv' in sel_cats),
                      ('optical', 'Оптика', 'optical' in sel_cats)],
        'roles':     [(role, role.pk in sel_role_ids) for role in all_roles],
    }

    # Build export URL (same params + export=xlsx)
    get_copy = request.GET.copy()
    get_copy['export'] = 'xlsx'
    export_url = '?' + get_copy.urlencode()

    return render(request, 'equipment_accounting/drone_stats.html', {
        'sections':      sections,
        'summary_cards': summary_cards,
        'total_all':     total_all,
        'filter_state':  filter_state,
        'is_filtered':   is_filtered,
        'export_url':    export_url,
    })


# ── Excel export ─────────────────────────────────────────────────────

EXPORT_COLS = [
    ('name',       'Назва'),
    ('count',      'В наявності'),
    ('unit',       'шт'),
    ('messenger',  'Для мессенджера'),
    ('new',        'Нові'),
    ('repair',     'В ремонті'),
    ('mfr_repair', 'В ремонті у виробника'),
    ('total',      'Підсумок'),
]

_SECTION_PRIORITY = {
    'День': 0, 'Ніч': 1, 'Оптика': 2,
    'Носій': 3, 'Мінувальник': 4, 'Перехоплювач': 5, 'Бомбардувальник': 6,
}

_SECTION_STYLE = {
    'День':   {'bg': 'BDD7EE', 'fg': '1F3864'},
    'Ніч':    {'bg': '2C3E6B', 'fg': 'FFFFFF'},
    'Оптика': {'bg': 'E2EFDA', 'fg': '375623'},
}
_DEFAULT_SECTION_STYLE = {'bg': 'F2F2F2', 'fg': '333333'}


@master_required
def uav_export_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    selected_cols = [c for c, _ in EXPORT_COLS if request.GET.get(f'col_{c}', '1') == '1']

    fpv_ct = ContentType.objects.get_for_model(FPVDroneType)
    opt_ct = ContentType.objects.get_for_model(OpticalDroneType)

    active_qs = UAVInstance.objects.exclude(status__in=['deleted', 'given']).select_related(
        'content_type', 'role'
    )

    fpv_pks = set(active_qs.filter(content_type=fpv_ct).values_list('object_id', flat=True))
    opt_pks = set(active_qs.filter(content_type=opt_ct).values_list('object_id', flat=True))

    fpv_type_map = (
        {dt.pk: dt for dt in FPVDroneType.objects.filter(pk__in=fpv_pks)
         .select_related('model', 'purpose', 'video_frequency')
         .prefetch_related('control_frequencies')} if fpv_pks else {}
    )
    opt_type_map = (
        {dt.pk: dt for dt in OpticalDroneType.objects.filter(pk__in=opt_pks)
         .select_related('model', 'purpose', 'video_template')
         .prefetch_related('control_frequencies')} if opt_pks else {}
    )

    sections = {}
    section_order = []

    for uav in active_qs:
        role_name = uav.role.name if uav.role_id else '—'

        if uav.content_type_id == fpv_ct.pk:
            dt = fpv_type_map.get(uav.object_id)
            if not dt:
                continue
            section_key = ('Ніч' if dt.has_thermal else 'День') if role_name == 'Ударні' else role_name
            type_label = _fmt_drone_type_name(dt, 'Радіо')
            type_key = ('fpv', dt.pk)
        elif uav.content_type_id == opt_ct.pk:
            dt = opt_type_map.get(uav.object_id)
            if not dt:
                continue
            section_key = 'Оптика'
            type_label = _fmt_drone_type_name(dt, 'Оптика')
            type_key = ('opt', dt.pk)
        else:
            continue

        if section_key not in sections:
            sections[section_key] = {'types': {}, 'type_order': []}
            section_order.append(section_key)

        if type_key not in sections[section_key]['types']:
            sections[section_key]['types'][type_key] = {
                'label': type_label,
                'ready': 0, 'inspection': 0, 'repair': 0, 'deferred': 0,
            }
            sections[section_key]['type_order'].append(type_key)

        t = sections[section_key]['types'][type_key]
        if uav.status in t:
            t[uav.status] += 1

    section_order.sort(key=lambda sk: (_SECTION_PRIORITY.get(sk, 99), sk))

    # Build workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'БПЛА'

    col_label_map = dict(EXPORT_COLS)
    col_map = {c: i + 1 for i, c in enumerate(selected_cols)}

    # Header row
    hdr_font = Font(bold=True, size=10, name='Calibri')
    hdr_fill = PatternFill(fill_type='solid', fgColor='D6DCE4')
    for col_key, col_idx in col_map.items():
        cell = ws.cell(row=1, column=col_idx, value=col_label_map[col_key])
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(wrap_text=True, vertical='center')
    ws.row_dimensions[1].height = 28

    row_num = 2
    for section_key in section_order:
        section = sections[section_key]
        style = _SECTION_STYLE.get(section_key, _DEFAULT_SECTION_STYLE)
        sec_fill = PatternFill(fill_type='solid', fgColor=style['bg'])
        sec_font = Font(bold=True, size=10, name='Calibri', color=style['fg'])

        # Section header row
        for col_key, col_idx in col_map.items():
            val = None
            if col_key == 'name':      val = section_key
            elif col_key == 'count':   val = 1
            elif col_key == 'messenger': val = section_key
            elif col_key == 'total':   val = 0
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = sec_font
            cell.fill = sec_fill
        row_num += 1

        # Drone type rows
        for type_key in section['type_order']:
            t = section['types'][type_key]
            count = t['ready'] + t['deferred']
            for col_key, col_idx in col_map.items():
                val = None
                if col_key == 'name':        val = t['label']
                elif col_key == 'count':     val = count
                elif col_key == 'unit':      val = 'шт'
                elif col_key == 'messenger': val = f"{t['label']} {count} шт" if count > 0 else None
                elif col_key == 'new':       val = t['inspection'] or None
                elif col_key == 'repair':    val = t['repair'] or None
                elif col_key == 'mfr_repair': val = None
                elif col_key == 'total':
                    c_count = col_map.get('count')
                    c_new = col_map.get('new')
                    c_repair = col_map.get('repair')
                    if c_count and c_new and c_repair:
                        val = f"={get_column_letter(c_count)}{row_num}+{get_column_letter(c_new)}{row_num}-{get_column_letter(c_repair)}{row_num}"
                    else:
                        val = count
                ws.cell(row=row_num, column=col_idx, value=val)
            row_num += 1

    # Column widths
    col_widths = {
        'name': 52, 'count': 13, 'unit': 6,
        'messenger': 58, 'new': 10, 'repair': 13,
        'mfr_repair': 24, 'total': 13,
    }
    for col_key, col_idx in col_map.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_key, 15)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="drones.xlsx"'
    return response


# ── UAV movement history ──────────────────────────────────────────────

def _fmt_freq(f):
    """Format a Frequency object as e.g. '900MHz' or '5.8GHz'."""
    v = f.value
    val_str = str(int(v)) if v == int(v) else str(v)
    return f"{val_str}{f.get_unit_display()}"


def _fmt_drone_type_name(dt, category):
    """Format drone type label.

    FPV  → 'ModelName (prop\') (freq1 freq2) (#purpose)'
    Optic → 'ModelName (prop\') distancekm [ніч]'
    """
    prop_str = f" ({dt.prop_size}')" if dt.prop_size else ''
    if category == 'Оптика':
        name = dt.model.name + prop_str
        if getattr(dt, 'video_template_id', None):
            name += f" {dt.video_template.max_distance}км"
        if dt.has_thermal:
            name += " ніч"
        return name

    # FPV / radio drone
    name = dt.model.name + prop_str
    ctrl_parts = [_fmt_freq(f) for f in sorted(dt.control_frequencies.all(), key=lambda f: f.value * 1000 if f.unit == 'ghz' else f.value)]
    video_part = _fmt_freq(dt.video_frequency) if getattr(dt, 'video_frequency_id', None) else ''
    if ctrl_parts or video_part:
        freqs_str = '-'.join(ctrl_parts)
        if video_part:
            freqs_str += (', ' if freqs_str else '') + video_part
        name += f"({freqs_str})"
    if dt.purpose_id:
        name += f" (#{dt.purpose.name})"
    return name


def _make_list_type_label(dt, is_opt):
    """Compact label for the equipment list table / badge grid.

    FPV   → 'ModelName (10") (900gh 5.8gh)'
    Optic → 'ModelName (10") 5км'
    """
    if not dt:
        return '—'
    name = dt.model.name
    if dt.prop_size:
        name += f' ({dt.prop_size}")'
    if is_opt:
        if getattr(dt, 'video_template_id', None):
            name += f' {dt.video_template.max_distance}км'
    else:
        ctrl_parts = [_fmt_freq(f) for f in sorted(dt.control_frequencies.all(), key=lambda f: f.value * 1000 if f.unit == 'ghz' else f.value)]
        video_part = _fmt_freq(dt.video_frequency) if getattr(dt, 'video_frequency_id', None) else ''
        if ctrl_parts or video_part:
            freqs_str = '-'.join(ctrl_parts)
            if video_part:
                freqs_str += (', ' if freqs_str else '') + video_part
            name += f'({freqs_str})'
    return name


def _build_role_groups(uav_objs, fpv_ct, opt_ct, fpv_types, opt_types):
    """Return hierarchical role_groups list.

    FPV drones are grouped by role (Ударні → День/Ніч sub-groups, others by
    role name).  Optical drones are appended as a single 'Оптика' section at
    the end, regardless of the UAV's assigned role.
    """
    fpv_rk_order = []
    fpv_rk_data = {}
    opt_type_order = []
    opt_type_data = {}

    for uav in uav_objs:
        if uav.content_type_id == fpv_ct.pk:
            dt = fpv_types.get(uav.object_id)
            role_name = uav.role.name if uav.role_id else '—'
            if role_name == 'Ударні' and dt is not None:
                sub_label = 'Ніч' if dt.has_thermal else 'День'
            else:
                sub_label = None
            rk = (role_name, sub_label)
            if rk not in fpv_rk_data:
                fpv_rk_data[rk] = {}
                fpv_rk_order.append(rk)
            type_key = dt.pk if dt else None
            if type_key not in fpv_rk_data[rk]:
                fpv_rk_data[rk][type_key] = {
                    'category': 'Радіо',
                    'type_label': _fmt_drone_type_name(dt, 'Радіо') if dt else '—',
                    'count': 0,
                }
            fpv_rk_data[rk][type_key]['count'] += 1

        elif uav.content_type_id == opt_ct.pk:
            dt = opt_types.get(uav.object_id)
            type_key = dt.pk if dt else None
            if type_key not in opt_type_data:
                opt_type_data[type_key] = {
                    'category': 'Оптика',
                    'type_label': _fmt_drone_type_name(dt, 'Оптика') if dt else '—',
                    'count': 0,
                }
                opt_type_order.append(type_key)
            opt_type_data[type_key]['count'] += 1

    role_groups = []
    for (role_name, sub_label) in fpv_rk_order:
        if not role_groups or role_groups[-1]['role_name'] != role_name:
            role_groups.append({'role_name': role_name, 'sub_groups': []})
        role_groups[-1]['sub_groups'].append({
            'sub_label': sub_label,
            'types': list(fpv_rk_data[(role_name, sub_label)].values()),
        })

    if opt_type_order:
        role_groups.append({
            'role_name': 'Оптика',
            'sub_groups': [{'sub_label': None, 'types': [opt_type_data[k] for k in opt_type_order]}],
        })

    return role_groups


@master_required
def uav_movements(request):
    """Show UAV movements grouped by calendar date; each date is expandable."""
    reason_filter = request.GET.get('reason', '')
    _loc_raw = request.GET.get('location', '')
    location_filter = int(_loc_raw) if _loc_raw.isdigit() else None
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    base_qs = UAVMovement.objects.select_related(
        'uav', 'uav__content_type', 'uav__role',
        'from_location', 'to_location',
        'moved_by', 'moved_by__profile',
    )
    if reason_filter:
        base_qs = base_qs.filter(reason=reason_filter)
    if location_filter:
        base_qs = base_qs.filter(
            Q(from_location_id=location_filter) | Q(to_location_id=location_filter)
        )
    if date_from:
        try:
            base_qs = base_qs.filter(created_at__date__gte=date.fromisoformat(date_from))
        except ValueError:
            pass
    if date_to:
        try:
            base_qs = base_qs.filter(created_at__date__lte=date.fromisoformat(date_to))
        except ValueError:
            pass

    # Outer group: by calendar date.  Inner group: by (reason, from, to, user).
    reason_labels = dict(UAVMovement.REASON_CHOICES)
    date_seen = {}   # date → date_group
    date_order = []  # ordered list of date keys

    for m in base_qs.order_by('-created_at'):
        d = m.created_at.date()
        if d not in date_seen:
            dg = {'date': m.created_at, 'count': 0, '_bseen': {}, '_border': []}
            date_seen[d] = dg
            date_order.append(d)
        dg = date_seen[d]
        dg['count'] += 1

        bk = (m.reason, m.from_location_id, m.to_location_id, m.moved_by_id)
        if bk not in dg['_bseen']:
            dg['_bseen'][bk] = {
                'reason': m.reason,
                'reason_label': reason_labels.get(m.reason, m.reason),
                'from_location_name': m.from_location.name if m.from_location else None,
                'to_location_name': m.to_location.name,
                'user_name': (m.moved_by.profile.display_name if hasattr(m.moved_by, 'profile') else m.moved_by.username) if m.moved_by else '—',
                'uav_objs': [],
                'movement_ids': [],
            }
            dg['_border'].append(bk)
        dg['_bseen'][bk]['uav_objs'].append(m.uav)
        dg['_bseen'][bk]['movement_ids'].append(m.pk)

    date_groups = []
    for d in date_order:
        dg = date_seen[d]
        bseen = dg.pop('_bseen')
        border = dg.pop('_border')
        dg['batches'] = [bseen[k] for k in border]
        date_groups.append(dg)

    # Paginate by date
    paginator = Paginator(date_groups, 30)
    page_obj = paginator.get_page(request.GET.get('page'))

    # Batch-fetch drone types for all UAVs on the current page
    fpv_ct = ContentType.objects.get_for_model(FPVDroneType)
    opt_ct = ContentType.objects.get_for_model(OpticalDroneType)
    fpv_pks, opt_pks = set(), set()
    for dg in page_obj.object_list:
        for batch in dg['batches']:
            for uav in batch['uav_objs']:
                if uav.content_type_id == fpv_ct.pk:
                    fpv_pks.add(uav.object_id)
                elif uav.content_type_id == opt_ct.pk:
                    opt_pks.add(uav.object_id)

    fpv_types = (
        {dt.pk: dt for dt in FPVDroneType.objects.filter(pk__in=fpv_pks)
         .select_related('model', 'power_template', 'purpose', 'video_frequency')
         .prefetch_related('control_frequencies')} if fpv_pks else {}
    )
    opt_types = (
        {dt.pk: dt for dt in OpticalDroneType.objects.filter(pk__in=opt_pks)
         .select_related('model', 'power_template', 'purpose', 'video_template')
         .prefetch_related('control_frequencies')} if opt_pks else {}
    )

    for dg in page_obj.object_list:
        for batch in dg['batches']:
            batch['role_groups'] = _build_role_groups(
                batch['uav_objs'], fpv_ct, opt_ct, fpv_types, opt_types
            )

    return render(request, 'equipment_accounting/uav_movements.html', {
        'page_obj': page_obj,
        'reason_filter': reason_filter,
        'location_filter': location_filter,
        'date_from': date_from,
        'date_to': date_to,
        'reason_choices': UAVMovement.REASON_CHOICES,
        'locations': Location.objects.all(),
    })


# ── UAV Detail / Assembly ────────────────────────────────────────────

def _component_matches_uav_template(component, uav):
    """Return True if the component is compatible with the UAV's drone type."""
    drone_type = uav.uav_type
    if component.kind == 'battery':
        return component.power_template_id == drone_type.power_template_id
    if component.kind == 'spool':
        return (uav.content_type.model == 'opticaldronetype'
                and component.video_template.drone_model_id == drone_type.video_template.drone_model_id
                and component.video_template.is_analog == drone_type.video_template.is_analog)
    return True  # other — no template restriction


@login_required
def uav_detail(request, pk):
    uav = get_object_or_404(UAVInstance, pk=pk)
    assigned_components = list(
        uav.components.select_related('power_template', 'video_template', 'other_type').all()
    )
    for comp in assigned_components:
        comp.type_display = str(comp)

    filled_kinds = {comp.kind for comp in assigned_components}
    drone_type = uav.uav_type

    free_qs_base = Component.objects.filter(
        assigned_to_uav=None, status__in=['in_use', 'disassembled']
    ).select_related('power_template', 'video_template', 'other_type')

    free_components = []
    if 'battery' not in filled_kinds:
        free_components += list(
            free_qs_base.filter(kind='battery', power_template_id=drone_type.power_template_id)
        )
    if uav.content_type.model == 'opticaldronetype' and 'spool' not in filled_kinds:
        free_components += list(
            free_qs_base.filter(
                kind='spool',
                video_template__drone_model_id=drone_type.video_template.drone_model_id,
                video_template__is_analog=drone_type.video_template.is_analog,
            )
        )

    for comp in free_components:
        if comp.kind == 'battery':
            comp.type_display = f"Батарея: {comp.power_template}"
        elif comp.kind == 'spool':
            drone_model = comp.video_template.drone_model if comp.video_template else None
            comp.type_display = f"Котушка: {comp.video_template} ({drone_model or '—'})"
        else:
            comp.type_display = str(comp.other_type)

    movements = uav.movements.select_related(
        'from_location', 'to_location', 'moved_by', 'moved_by__profile'
    ).order_by('-created_at')

    kit_status = uav.get_kit_status()
    photos = uav.photos.all()
    pending_movement = None
    if uav.status == 'transit':
        pending_movement = uav.movements.filter(confirmed_at__isnull=True).order_by('-created_at').first()
    return render(request, 'equipment_accounting/uav_detail.html', {
        'uav': uav,
        'assigned_components': assigned_components,
        'free_components': free_components,
        'kit_status': kit_status,
        'movements': movements,
        'pending_movement': pending_movement,
        'locations': Location.objects.all(),
        'photos': photos,
        'can_edit_uav': _can(request.user, PERM_CHANGE_UAV),
    })


@login_required
def movement_batch_delete(request):
    if not request.user.is_superuser:
        raise PermissionDenied
    if request.method == 'POST':
        ids = request.POST.getlist('movement_ids')
        deleted, _ = UAVMovement.objects.filter(pk__in=ids).delete()
        messages.success(request, f'Видалено {deleted} переміщень.')
    return redirect('equipment_accounting:uav_movements')


@uav_perm_required(PERM_CHANGE_UAV)
def uav_photo_upload(request, uav_pk):
    uav = get_object_or_404(UAVInstance, pk=uav_pk)
    if request.method == 'POST':
        files = request.FILES.getlist('photos')
        for f in files:
            UAVPhoto.objects.create(uav=uav, image=f)
    return redirect(reverse('equipment_accounting:uav_detail', args=[uav_pk]))


@uav_perm_required(PERM_CHANGE_UAV)
def uav_photo_delete(request, photo_pk):
    photo = get_object_or_404(UAVPhoto, pk=photo_pk)
    uav_pk = photo.uav_id
    if request.method == 'POST':
        photo.image.delete(save=False)
        photo.delete()
    return redirect(reverse('equipment_accounting:uav_detail', args=[uav_pk]))


@uav_perm_required(PERM_CHANGE_UAV)
def uav_photo_edit(request, photo_pk):
    photo = get_object_or_404(UAVPhoto, pk=photo_pk)
    uav_pk = photo.uav_id
    if request.method == 'POST':
        caption = request.POST.get('caption', '').strip()
        photo.caption = caption
        photo.save(update_fields=['caption'])
    return redirect(reverse('equipment_accounting:uav_detail', args=[uav_pk]))


@uav_perm_required(PERM_CHANGE_UAV)
def uav_move(request, pk):
    """Initiate transit: set UAV to 'transit' and record an unconfirmed movement."""
    uav = get_object_or_404(UAVInstance, pk=pk)
    if request.method != 'POST':
        return redirect(reverse('equipment_accounting:uav_detail', args=[pk]))

    if uav.status not in UAVInstance.ACTIVE_STATUSES or uav.status == 'transit':
        messages.error(request, 'Неможливо перемістити дрон у поточному статусі.')
        return redirect(reverse('equipment_accounting:uav_detail', args=[pk]))

    to_location_id = request.POST.get('to_location_id')
    to_location = Location.objects.filter(pk=to_location_id).first() if to_location_id else None
    if not to_location:
        messages.error(request, 'Оберіть локацію призначення.')
        return redirect(reverse('equipment_accounting:uav_detail', args=[pk]))

    if to_location == uav.current_location:
        messages.error(request, 'Дрон вже знаходиться на цій локації.')
        return redirect(reverse('equipment_accounting:uav_detail', args=[pk]))

    notes = request.POST.get('notes', '').strip()
    UAVMovement.objects.create(
        uav=uav,
        from_location=uav.current_location,
        to_location=to_location,
        moved_by=request.user,
        reason='transferred',
        notes=notes,
        pre_transit_status=uav.status,
    )
    uav.status = 'transit'
    uav.pending_to_location = to_location
    uav.save(update_fields=['status', 'pending_to_location', 'updated_at'])

    messages.success(request, f'БПЛА відправлено до "{to_location.name}". Очікується підтвердження прибуття.')
    return redirect(reverse('equipment_accounting:uav_detail', args=[pk]))


@uav_perm_required(PERM_CHANGE_UAV)
def uav_confirm_arrival(request, movement_pk):
    """Confirm arrival: update current_location and restore pre-transit status."""
    from django.utils import timezone as tz
    movement = get_object_or_404(UAVMovement, pk=movement_pk)
    uav = movement.uav
    if request.method != 'POST':
        return redirect(reverse('equipment_accounting:uav_detail', args=[uav.pk]))

    if movement.confirmed_at is not None:
        messages.warning(request, 'Прибуття вже підтверджено.')
        return redirect(reverse('equipment_accounting:uav_detail', args=[uav.pk]))

    movement.confirmed_at = tz.now()
    movement.confirmed_by = request.user
    movement.save(update_fields=['confirmed_at', 'confirmed_by'])

    uav.current_location = movement.to_location
    uav.pending_to_location = None
    uav.status = movement.pre_transit_status or 'inspection'
    uav.save(update_fields=['current_location', 'pending_to_location', 'status', 'updated_at'])

    messages.success(request, f'Прибуття БПЛА до "{movement.to_location.name}" підтверджено.')
    return redirect(reverse('equipment_accounting:uav_detail', args=[uav.pk]))


@uav_perm_required(PERM_CHANGE_UAV)
def uav_attach_component(request, uav_pk, component_pk):
    if request.method != 'POST':
        return redirect('equipment_accounting:uav_detail', pk=uav_pk)
    uav = get_object_or_404(UAVInstance, pk=uav_pk)
    component = get_object_or_404(Component, pk=component_pk)
    if component.assigned_to_uav is not None:
        messages.error(request, 'Комплектуюча вже закріплена за іншим БПЛА.')
    elif not _component_matches_uav_template(component, uav):
        messages.error(request, 'Ця комплектуюча не сумісна з даним типом БПЛА.')
    elif component.kind in ('battery', 'spool') and uav.components.filter(kind=component.kind).exists():
        messages.error(request, 'БПЛА вже має комплектуючу цього типу.')
    elif not _component_matches_uav_template(component, uav):
        messages.error(request, 'Комплектуюча не підходить за шаблоном до цього БПЛА.')
    else:
        component.assigned_to_uav = uav
        component.status = 'in_use'
        component.save(update_fields=['assigned_to_uav', 'status', 'updated_at'])
        messages.success(request, 'Комплектуючу закріплено.')
    return redirect('equipment_accounting:uav_detail', pk=uav_pk)


@uav_perm_required(PERM_CHANGE_UAV)
def uav_detach_component(request, uav_pk, component_pk):
    if request.method != 'POST':
        return redirect('equipment_accounting:uav_detail', pk=uav_pk)
    uav = get_object_or_404(UAVInstance, pk=uav_pk)
    component = get_object_or_404(Component, pk=component_pk)
    if component.assigned_to_uav_id != uav.pk:
        messages.error(request, 'Комплектуюча не закріплена за цим БПЛА.')
    else:
        component.assigned_to_uav = None
        component.status = 'disassembled'
        component.save(update_fields=['assigned_to_uav', 'status', 'updated_at'])
        messages.success(request, 'Комплектуючу відкріплено.')
    return redirect('equipment_accounting:uav_detail', pk=uav_pk)


# ── Bulk actions ────────────────────────────────────────────────────

@uav_perm_required(PERM_CHANGE_UAV)
def uav_toggle_given(request, pk):
    """Give away a ready UAV or return a given UAV to the workshop."""
    if request.method != "POST":
        return redirect(_list_url("drones"))
    uav = get_object_or_404(UAVInstance, pk=pk)
    workshop = Location.objects.filter(name='Майстерня').first()
    next_url = request.POST.get('next') or _list_url("drones")

    if uav.status == 'given':
        # Return via transit to workshop
        prev_location = uav.current_location
        UAVMovement.objects.create(
            uav=uav,
            from_location=prev_location,
            to_location=workshop,
            moved_by=request.user,
            reason='returned',
            pre_transit_status='inspection',
        )
        uav.status = 'transit'
        uav.pending_to_location = workshop
        uav.position = None
        uav.save(update_fields=['status', 'pending_to_location', 'position', 'updated_at'])
    elif uav.status == 'ready':
        to_location_id = request.POST.get('to_location_id')
        to_location = Location.objects.filter(pk=to_location_id).first() if to_location_id else None
        position = None
        if to_location and to_location.name == 'Позиція':
            position_id = request.POST.get('position_id')
            position_name_new = request.POST.get('position_name_new', '').strip()
            if position_id:
                position = Position.objects.filter(pk=position_id).first()
            elif position_name_new:
                position, _ = Position.objects.get_or_create(name=position_name_new)
        prev_location = uav.current_location
        if to_location:
            # Send via transit; status becomes 'given' after arrival confirmation
            UAVMovement.objects.create(
                uav=uav,
                from_location=prev_location,
                to_location=to_location,
                moved_by=request.user,
                reason='given',
                pre_transit_status='given',
            )
            uav.status = 'transit'
            uav.pending_to_location = to_location
            uav.position = position
            uav.save(update_fields=['status', 'pending_to_location', 'position', 'updated_at'])
        else:
            uav.status = 'given'
            uav.save(update_fields=['status', 'updated_at'])
    else:
        messages.error(request, 'Віддати можна лише готовий дрон.')
    return redirect(next_url)


@uav_perm_required(PERM_CHANGE_UAV)
def uav_bulk_action(request):
    """Handle bulk status change or bulk delete for selected UAVs."""
    if request.method != "POST":
        return redirect("equipment_accounting:equipment_list")

    ids = request.POST.getlist("selected")
    action = request.POST.get("bulk_action", "")

    if not ids:
        messages.warning(request, "Нічого не обрано.")
        return redirect("equipment_accounting:equipment_list")

    qs = UAVInstance.objects.filter(pk__in=ids)
    count = qs.count()

    to_location_id = request.POST.get('to_location_id')
    to_location = Location.objects.filter(pk=to_location_id).first() if to_location_id else None
    position = None
    if to_location and to_location.name == 'Позиція':
        position_id = request.POST.get('position_id')
        position_name_new = request.POST.get('position_name_new', '').strip()
        if position_id:
            position = Position.objects.filter(pk=position_id).first()
        elif position_name_new:
            position, _ = Position.objects.get_or_create(name=position_name_new)

    if action == "delete":
        if not _can(request.user, PERM_DELETE_UAV):
            raise PermissionDenied
        qs.update(status='deleted')
        messages.success(request, f"Видалено {count} БПЛА.")
    elif action == "given":
        eligible = qs.filter(status='ready')
        given_count = eligible.count()
        skipped = count - given_count
        for uav in eligible.select_related('current_location'):
            prev = uav.current_location
            if to_location:
                UAVMovement.objects.create(
                    uav=uav,
                    from_location=prev,
                    to_location=to_location,
                    moved_by=request.user,
                    reason='given',
                    pre_transit_status='given',
                )
                uav.status = 'transit'
                uav.pending_to_location = to_location
                uav.position = position
                uav.save(update_fields=['status', 'pending_to_location', 'position', 'updated_at'])
            else:
                uav.status = 'given'
                uav.save(update_fields=['status', 'updated_at'])
        msg = f"Віддано {given_count} БПЛА разом з комплектуючими."
        if skipped:
            msg += f" Пропущено {skipped} (не готові)."
        messages.success(request, msg)
    elif action == 'repair':
        prev_statuses = {uav.pk: uav.current_location for uav in qs.select_related('current_location')}
        qs.update(status='repair')
        if to_location:
            for uav in qs:
                uav.current_location = to_location
                uav.save(update_fields=['current_location', 'updated_at'])
                UAVMovement.objects.create(
                    uav=uav,
                    from_location=prev_statuses.get(uav.pk),
                    to_location=to_location,
                    moved_by=request.user,
                    reason='repair',
                )
        messages.success(request, f"Статус {count} БПЛА змінено на \"Ремонт\".")
    elif action in dict(UAVInstance.STATUS_CHOICES):
        qs.update(status=action)
        label = dict(UAVInstance.STATUS_CHOICES)[action]
        messages.success(request, f"Статус {count} БПЛА змінено на \"{label}\".")
    else:
        messages.error(request, "Невідома дія.")

    return redirect("equipment_accounting:equipment_list")


# ── UAV CRUD ────────────────────────────────────────────────────────

def _create_kit_components(uav, drone_type_obj, with_battery=True, with_spool=True):
    """Create battery and/or spool components for a UAV based on its drone type."""
    if with_battery:
        Component.objects.create(
            kind='battery',
            power_template=drone_type_obj.power_template,
            status='in_use',
            assigned_to_uav=uav,
        )
    if with_spool and isinstance(drone_type_obj, OpticalDroneType):
        Component.objects.create(
            kind='spool',
            video_template=drone_type_obj.video_template,
            status='in_use',
            assigned_to_uav=uav,
        )


def _build_drone_types_kit_data():
    """Return JSON-serialisable dict mapping 'ct_id-obj_id' to kit info."""
    import json
    data = {}
    fpv_ct = ContentType.objects.get_for_model(FPVDroneType)
    for dt in FPVDroneType.objects.select_related('power_template', 'model', 'model__manufacturer'):
        data[f"{fpv_ct.pk}-{dt.pk}"] = {
            "category": "fpv",
            "battery_name": str(dt.power_template),
            "spool_name": None,
        }
    opt_ct = ContentType.objects.get_for_model(OpticalDroneType)
    for dt in OpticalDroneType.objects.select_related(
        'power_template', 'video_template', 'model', 'model__manufacturer'
    ):
        data[f"{opt_ct.pk}-{dt.pk}"] = {
            "category": "optical",
            "battery_name": str(dt.power_template),
            "spool_name": str(dt.video_template),
        }
    return json.dumps(data)


@uav_perm_required(PERM_ADD_UAV)
def uav_create(request):
    workshop = Location.objects.filter(name='Майстерня').first()

    if request.method == "POST":
        form = UAVInstanceForm(request.POST)
        if form.is_valid():
            quantity      = form.cleaned_data.get("quantity", 1)
            with_battery  = form.cleaned_data.get("with_battery", True)
            with_spool    = form.cleaned_data.get("with_spool", True)
            from_location = form.cleaned_data.get("from_location")
            role          = form.cleaned_data.get("role")
            ct_id, obj_id = form.cleaned_data["drone_type"].split("-")
            ct = ContentType.objects.get(pk=int(ct_id))
            drone_type_obj = ct.get_object_for_this_type(pk=int(obj_id))
            created = []
            for _ in range(quantity):
                uav = UAVInstance.objects.create(
                    content_type_id=int(ct_id),
                    object_id=int(obj_id),
                    status="inspection",
                    created_by=request.user,
                    current_location=workshop,
                    role=role,
                )
                if with_battery or with_spool:
                    _create_kit_components(
                        uav, drone_type_obj,
                        with_battery=with_battery,
                        with_spool=with_spool,
                    )
                # Record movement: from_location → workshop (skip if workshop not configured)
                if workshop:
                    UAVMovement.objects.create(
                        uav=uav,
                        from_location=from_location,
                        to_location=workshop,
                        moved_by=request.user,
                        reason='created',
                    )
                created.append(uav)
            msg = f"Додано {quantity} БПЛА." if quantity > 1 else "БПЛА додано."
            messages.success(request, msg)
            return redirect("equipment_accounting:equipment_list")
    else:
        form = UAVInstanceForm()

    return render(request, "equipment_accounting/uav_create_form.html", {
        "form": form,
        "title": "Додати БПЛА",
        "drone_types_kit_json": _build_drone_types_kit_data(),
        "locations": Location.objects.all(),
    })


@uav_perm_required(PERM_CHANGE_UAV)
def uav_edit(request, pk):
    uav = get_object_or_404(UAVInstance, pk=pk)
    if request.method == "POST":
        form = UAVInstanceForm(request.POST, instance=uav)
        if form.is_valid():
            form.save()
            messages.success(request, "БПЛА оновлено.")
            return redirect("equipment_accounting:equipment_list")
    else:
        form = UAVInstanceForm(instance=uav)
    return render(request, "equipment_accounting/equipment_form.html", {
        "form": form, "title": "Редагувати БПЛА",
    })


@uav_perm_required(PERM_DELETE_UAV)
def uav_delete(request, pk):
    uav = get_object_or_404(UAVInstance, pk=pk)
    components = list(uav.components.select_related('content_type').all())
    if request.method == "POST":
        delete_components = request.POST.get('delete_components') == '1'
        if components:
            if delete_components:
                uav.components.all().delete()
            else:
                uav.components.all().update(assigned_to_uav=None, status='disassembled')
        uav.status = 'deleted'
        uav.save(update_fields=['status', 'updated_at'])
        messages.success(request, "БПЛА видалено.")
        return redirect("equipment_accounting:equipment_list")
    for comp in components:
        comp.type_display = str(comp)
    return render(request, "equipment_accounting/equipment_confirm_delete.html", {
        "object": uav, "title": "Видалити БПЛА",
        "cancel_url": _list_url("drones"),
        "uav_components": components,
    })


# ── Manufacturer CRUD ───────────────────────────────────────────────

@uav_perm_required(PERM_ADD_MANUFACTURER)
def manufacturer_create(request):
    if request.method == "POST":
        form = ManufacturerForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Виробника додано.")
            return redirect(_list_url("types"))
    else:
        form = ManufacturerForm()
    return render(request, "equipment_accounting/equipment_form.html", {
        "form": form, "title": "Додати виробника", "tab_redirect": "types",
    })


@uav_perm_required(PERM_CHANGE_MANUFACTURER)
def manufacturer_edit(request, pk):
    manufacturer = get_object_or_404(Manufacturer, pk=pk)
    if request.method == "POST":
        form = ManufacturerForm(request.POST, instance=manufacturer)
        if form.is_valid():
            form.save()
            messages.success(request, "Виробника оновлено.")
            return redirect(_list_url("types"))
    else:
        form = ManufacturerForm(instance=manufacturer)
    return render(request, "equipment_accounting/equipment_form.html", {
        "form": form, "title": "Редагувати виробника", "tab_redirect": "types",
    })


@uav_perm_required(PERM_DELETE_MANUFACTURER)
def manufacturer_delete(request, pk):
    manufacturer = get_object_or_404(Manufacturer, pk=pk)
    if request.method == "POST":
        manufacturer.delete()
        messages.success(request, "Виробника видалено.")
        return redirect(_list_url("types"))
    return render(request, "equipment_accounting/equipment_confirm_delete.html", {
        "object": manufacturer, "title": "Видалити виробника",
        "cancel_url": _list_url("types"),
    })


# ── DroneModel CRUD ─────────────────────────────────────────────────

@uav_perm_required(PERM_ADD_DRONEMODEL)
def drone_model_create(request):
    if request.method == "POST":
        form = DroneModelForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Модель дрона додано.")
            return redirect(_list_url("types"))
    else:
        form = DroneModelForm()
    return render(request, "equipment_accounting/equipment_form.html", {
        "form": form, "title": "Додати модель дрона", "tab_redirect": "types",
    })


@uav_perm_required(PERM_CHANGE_DRONEMODEL)
def drone_model_edit(request, pk):
    drone_model = get_object_or_404(DroneModel, pk=pk)
    if request.method == "POST":
        form = DroneModelForm(request.POST, instance=drone_model)
        if form.is_valid():
            form.save()
            messages.success(request, "Модель дрона оновлено.")
            return redirect(_list_url("types"))
    else:
        form = DroneModelForm(instance=drone_model)
    return render(request, "equipment_accounting/equipment_form.html", {
        "form": form, "title": "Редагувати модель дрона", "tab_redirect": "types",
    })


@uav_perm_required(PERM_DELETE_DRONEMODEL)
def drone_model_delete(request, pk):
    drone_model = get_object_or_404(DroneModel, pk=pk)
    if request.method == "POST":
        drone_model.delete()
        messages.success(request, "Модель дрона видалено.")
        return redirect(_list_url("types"))
    return render(request, "equipment_accounting/equipment_confirm_delete.html", {
        "object": drone_model, "title": "Видалити модель дрона",
        "cancel_url": _list_url("types"),
    })


# ── FPV Drone Type CRUD ─────────────────────────────────────────────

@uav_perm_required(PERM_ADD_FPVTYPE)
def fpv_type_create(request):
    if request.method == "POST":
        form = FPVDroneTypeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Тип радіо дрона додано.")
            return redirect(_list_url("types"))
    else:
        form = FPVDroneTypeForm()
    return render(request, "equipment_accounting/equipment_form.html", {
        "form": form, "title": "Додати тип радіо дрона", "tab_redirect": "types",
    })


@uav_perm_required(PERM_CHANGE_FPVTYPE)
def fpv_type_edit(request, pk):
    drone_type = get_object_or_404(FPVDroneType, pk=pk)
    if request.method == "POST":
        form = FPVDroneTypeForm(request.POST, instance=drone_type)
        if form.is_valid():
            form.save()
            messages.success(request, "Тип радіо дрона оновлено.")
            return redirect(_list_url("types"))
    else:
        form = FPVDroneTypeForm(instance=drone_type)
    return render(request, "equipment_accounting/equipment_form.html", {
        "form": form, "title": "Редагувати тип радіо дрона", "tab_redirect": "types",
    })


@uav_perm_required(PERM_DELETE_FPVTYPE)
def fpv_type_delete(request, pk):
    drone_type = get_object_or_404(FPVDroneType, pk=pk)
    if request.method == "POST":
        drone_type.delete()
        messages.success(request, "Тип радіо дрона видалено.")
        return redirect(_list_url("types"))
    return render(request, "equipment_accounting/equipment_confirm_delete.html", {
        "object": drone_type, "title": "Видалити тип радіо дрона",
        "cancel_url": _list_url("types"),
    })


# ── Optical Drone Type CRUD ────────────────────────────────────────

@uav_perm_required(PERM_ADD_OPTICALTYPE)
def optical_type_create(request):
    if request.method == "POST":
        form = OpticalDroneTypeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Тип оптичного дрона додано.")
            return redirect(_list_url("types"))
    else:
        form = OpticalDroneTypeForm()
    return render(request, "equipment_accounting/equipment_form.html", {
        "form": form, "title": "Додати тип оптичного дрона", "tab_redirect": "types",
    })


@uav_perm_required(PERM_CHANGE_OPTICALTYPE)
def optical_type_edit(request, pk):
    drone_type = get_object_or_404(OpticalDroneType, pk=pk)
    if request.method == "POST":
        form = OpticalDroneTypeForm(request.POST, instance=drone_type)
        if form.is_valid():
            form.save()
            messages.success(request, "Тип оптичного дрона оновлено.")
            return redirect(_list_url("types"))
    else:
        form = OpticalDroneTypeForm(instance=drone_type)
    return render(request, "equipment_accounting/equipment_form.html", {
        "form": form, "title": "Редагувати тип оптичного дрона", "tab_redirect": "types",
    })


@uav_perm_required(PERM_DELETE_OPTICALTYPE)
def optical_type_delete(request, pk):
    drone_type = get_object_or_404(OpticalDroneType, pk=pk)
    if request.method == "POST":
        drone_type.delete()
        messages.success(request, "Тип оптичного дрона видалено.")
        return redirect(_list_url("types"))
    return render(request, "equipment_accounting/equipment_confirm_delete.html", {
        "object": drone_type, "title": "Видалити тип оптичного дрона",
        "cancel_url": _list_url("types"),
    })


# ── Component CRUD ──────────────────────────────────────────────────

_COMPONENT_EXTRA = lambda: {
    "uav_filter_url": reverse("equipment_accounting:component_available_uavs")
}


@login_required
def component_available_uavs(request):
    """Return JSON list of UAVs available for a given component kind and template."""
    kind = request.GET.get("kind", "")
    power_template_id = request.GET.get("power_template") or None
    video_template_id = request.GET.get("video_template") or None
    exclude_pk = None
    try:
        exclude_pk = int(request.GET.get("exclude", ""))
    except (ValueError, TypeError):
        pass
    uavs = _get_available_uavs_for_kind(
        kind,
        exclude_component_pk=exclude_pk,
        power_template_id=power_template_id,
        video_template_id=video_template_id,
    )
    return JsonResponse({
        "uavs": [{"id": u.pk, "text": str(u)} for u in uavs]
    })


@uav_perm_required(PERM_CHANGE_COMPONENT)
def component_bulk_action(request):
    """Handle bulk actions (mark damaged / restore / delete) for selected components."""
    if request.method != "POST":
        return redirect("equipment_accounting:equipment_list")

    ids = request.POST.getlist("selected")
    action = request.POST.get("bulk_action", "")
    redirect_url = f"{reverse('equipment_accounting:equipment_list')}?tab=components"

    if not ids:
        messages.warning(request, "Нічого не обрано.")
        return redirect(redirect_url)

    qs = Component.objects.filter(pk__in=ids)
    count = qs.count()

    if action == "damaged":
        qs.update(status="damaged", assigned_to_uav_id=None)
        messages.success(request, f"Позначено пошкодженими: {count}.")
    elif action == "restore":
        qs.update(status="in_use")
        messages.success(request, f"Відновлено: {count}.")
    elif action == "delete":
        if not request.user.has_perm(PERM_DELETE_COMPONENT):
            raise PermissionDenied
        qs.delete()
        messages.success(request, f"Видалено {count} комплектуючих.")
    else:
        messages.warning(request, "Оберіть дію.")

    return redirect(redirect_url)


@uav_perm_required(PERM_ADD_COMPONENT)
def component_create(request):
    if request.method == "POST":
        form = ComponentForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Комплектуючу додано.")
            return redirect(_list_url("components"))
    else:
        form = ComponentForm()
    return render(request, "equipment_accounting/equipment_form.html", {
        "form": form, "title": "Додати комплектуючу", "tab_redirect": "components",
        **_COMPONENT_EXTRA(),
    })


@uav_perm_required(PERM_CHANGE_COMPONENT)
def component_edit(request, pk):
    component = get_object_or_404(Component, pk=pk)
    if request.method == "POST":
        form = ComponentForm(request.POST, instance=component)
        if form.is_valid():
            form.save()
            messages.success(request, "Комплектуючу оновлено.")
            return redirect(_list_url("components"))
    else:
        form = ComponentForm(instance=component)
    return render(request, "equipment_accounting/equipment_form.html", {
        "form": form, "title": "Редагувати комплектуючу", "tab_redirect": "components",
        **_COMPONENT_EXTRA(),
    })


@uav_perm_required(PERM_CHANGE_COMPONENT)
def component_mark_damaged(request, pk):
    """Mark a component as damaged and detach it from any UAV."""
    if request.method != "POST":
        return redirect(_list_url("components"))
    component = get_object_or_404(Component, pk=pk)
    component.status = "damaged"
    component.assigned_to_uav = None
    component.save(update_fields=["status", "assigned_to_uav", "updated_at"])
    messages.success(request, "Комплектуючу позначено як пошкоджену.")
    next_url = request.POST.get("next") or _list_url("components")
    return redirect(next_url)


@uav_perm_required(PERM_CHANGE_COMPONENT)
def component_restore(request, pk):
    """Restore a damaged or disassembled component to in_use status."""
    if request.method != "POST":
        return redirect(_list_url("components"))
    component = get_object_or_404(Component, pk=pk)
    component.status = "in_use"
    component.save(update_fields=["status", "updated_at"])
    messages.success(request, "Комплектуючу відновлено.")
    next_url = request.POST.get("next") or _list_url("components")
    return redirect(next_url)


@uav_perm_required(PERM_DELETE_COMPONENT)
def component_delete(request, pk):
    component = get_object_or_404(Component, pk=pk)
    if request.method == "POST":
        component.delete()
        messages.success(request, "Комплектуючу видалено.")
        return redirect(_list_url("components"))
    return render(request, "equipment_accounting/equipment_confirm_delete.html", {
        "object": component, "title": "Видалити комплектуючу",
        "cancel_url": _list_url("components"),
    })


# ── Location CRUD ───────────────────────────────────────────────────

@uav_perm_required(PERM_ADD_LOCATION)
def location_create(request):
    if request.method == "POST":
        form = LocationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Локацію додано.")
            return redirect(_list_url("locations"))
    else:
        form = LocationForm()
    return render(request, "equipment_accounting/equipment_form.html", {
        "form": form, "title": "Додати локацію", "tab_redirect": "locations",
    })


@uav_perm_required(PERM_CHANGE_LOCATION)
def location_edit(request, pk):
    location = get_object_or_404(Location, pk=pk)
    if request.method == "POST":
        form = LocationForm(request.POST, instance=location)
        if form.is_valid():
            form.save()
            messages.success(request, "Локацію оновлено.")
            return redirect(_list_url("locations"))
    else:
        form = LocationForm(instance=location)
    return render(request, "equipment_accounting/equipment_form.html", {
        "form": form, "title": "Редагувати локацію", "tab_redirect": "locations",
    })


@uav_perm_required(PERM_DELETE_LOCATION)
def location_delete(request, pk):
    location = get_object_or_404(Location, pk=pk)
    # Prevent deleting locations that have UAVs assigned
    uav_count = location.current_uavs.exclude(status='deleted').count()
    if request.method == "POST":
        if uav_count:
            messages.error(request, f"Неможливо видалити: {uav_count} БПЛА перебуває на цій локації.")
            return redirect(_list_url("locations"))
        try:
            location.delete()
            messages.success(request, "Локацію видалено.")
        except ProtectedError:
            messages.error(request, "Неможливо видалити: локація має пов'язані переміщення БПЛА.")
        return redirect(_list_url("locations"))
    return render(request, "equipment_accounting/equipment_confirm_delete.html", {
        "object": location, "title": "Видалити локацію",
        "cancel_url": _list_url("locations"),
        "extra_warning": f"На локації є {uav_count} активних БПЛА." if uav_count else "",
    })


# ── Position CRUD ───────────────────────────────────────────────────

@uav_perm_required(PERM_ADD_POSITION)
def position_create(request):
    if request.method == "POST":
        form = PositionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Позицію додано.")
            return redirect(_list_url("locations"))
    else:
        form = PositionForm()
    return render(request, "equipment_accounting/equipment_form.html", {
        "form": form, "title": "Додати позицію", "tab_redirect": "locations",
    })


@uav_perm_required(PERM_CHANGE_POSITION)
def position_edit(request, pk):
    position = get_object_or_404(Position, pk=pk)
    if request.method == "POST":
        form = PositionForm(request.POST, instance=position)
        if form.is_valid():
            form.save()
            messages.success(request, "Позицію оновлено.")
            return redirect(_list_url("locations"))
    else:
        form = PositionForm(instance=position)
    return render(request, "equipment_accounting/equipment_form.html", {
        "form": form, "title": "Редагувати позицію", "tab_redirect": "locations",
    })


@uav_perm_required(PERM_DELETE_POSITION)
def position_delete(request, pk):
    position = get_object_or_404(Position, pk=pk)
    uav_count = position.uavs.count()
    if request.method == "POST":
        if uav_count:
            messages.error(request, f"Неможливо видалити: {uav_count} БПЛА має цю позицію.")
            return redirect(_list_url("locations"))
        position.delete()
        messages.success(request, "Позицію видалено.")
        return redirect(_list_url("locations"))
    return render(request, "equipment_accounting/equipment_confirm_delete.html", {
        "object": position, "title": "Видалити позицію",
        "cancel_url": _list_url("locations"),
        "extra_warning": f"На позиції є {uav_count} активних БПЛА." if uav_count else "",
    })


# ── PowerTemplate CRUD ──────────────────────────────────────────────

@uav_perm_required(PERM_ADD_POWERTEMPLATE)
def power_template_create(request):
    if request.method == "POST":
        form = PowerTemplateForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Шаблон живлення додано.")
            return redirect(_list_url("templates"))
    else:
        form = PowerTemplateForm()
    return render(request, "equipment_accounting/equipment_form.html", {
        "form": form, "title": "Додати шаблон живлення", "tab_redirect": "templates",
    })


@uav_perm_required(PERM_CHANGE_POWERTEMPLATE)
def power_template_edit(request, pk):
    template = get_object_or_404(PowerTemplate, pk=pk, is_deleted=False)
    if request.method == "POST":
        form = PowerTemplateForm(request.POST, instance=template)
        if form.is_valid():
            form.save()
            messages.success(request, "Шаблон живлення оновлено.")
            return redirect(_list_url("templates"))
    else:
        form = PowerTemplateForm(instance=template)
    return render(request, "equipment_accounting/equipment_form.html", {
        "form": form, "title": "Редагувати шаблон живлення", "tab_redirect": "templates",
    })


@uav_perm_required(PERM_DELETE_POWERTEMPLATE)
def power_template_delete(request, pk):
    messages.error(request, "Видалення шаблонів живлення заборонено.")
    return redirect(_list_url("templates"))


# ── VideoTemplate CRUD ──────────────────────────────────────────────

@uav_perm_required(PERM_ADD_VIDEOTEMPLATE)
def video_template_create(request):
    if request.method == "POST":
        form = VideoTemplateForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Шаблон відео додано.")
            return redirect(_list_url("templates"))
    else:
        form = VideoTemplateForm()
    return render(request, "equipment_accounting/equipment_form.html", {
        "form": form, "title": "Додати шаблон відео", "tab_redirect": "templates",
    })


@uav_perm_required(PERM_CHANGE_VIDEOTEMPLATE)
def video_template_edit(request, pk):
    template = get_object_or_404(VideoTemplate, pk=pk, is_deleted=False)
    if request.method == "POST":
        form = VideoTemplateForm(request.POST, instance=template)
        if form.is_valid():
            form.save()
            messages.success(request, "Шаблон відео оновлено.")
            return redirect(_list_url("templates"))
    else:
        form = VideoTemplateForm(instance=template)
    return render(request, "equipment_accounting/equipment_form.html", {
        "form": form, "title": "Редагувати шаблон відео", "tab_redirect": "templates",
    })


@uav_perm_required(PERM_DELETE_VIDEOTEMPLATE)
def video_template_delete(request, pk):
    messages.error(request, "Видалення шаблонів відео заборонено.")
    return redirect(_list_url("templates"))
